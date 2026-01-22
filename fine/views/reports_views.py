# fine/views/reports_views.py
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET, require_POST
from django.db.models import Sum, Count, Avg, F, Q
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ..models import (
    Income, Expense, Purchase, Payroll,
    Attendance, AttendanceSummary, Cat
)


def reports(request):
    """Main reports page view"""
    today = timezone.now().date()
    start_of_month = today.replace(day=1)
    last_day = (today.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    context = {
        'today_date': today,
        'start_date': start_of_month,
        'end_date': last_day,
    }
    return render(request, 'reports.html', context)


@require_GET
def get_report_data(request):
    """Unified API endpoint to fetch report data for all report types"""
    try:
        report_type = request.GET.get('type', 'all')
        period_type = request.GET.get('period', 'daily')
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None

        if not start_date or not end_date:
            today = timezone.now().date()
            if period_type == 'daily':
                start_date = end_date = today
            elif period_type == 'weekly':
                start_date = today - timedelta(days=today.weekday())
                end_date = start_date + timedelta(days=6)
            elif period_type == 'monthly':
                start_date = today.replace(day=1)
                next_month = start_date.replace(
                    month=start_date.month % 12 + 1,
                    year=start_date.year + (start_date.month // 12)
                )
                end_date = next_month - timedelta(days=1)
            else:
                start_date = today.replace(month=1, day=1)
                end_date = today.replace(month=12, day=31)

        response_data = {
            'success': True,
            'period_type': period_type,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'data': {}
        }

        if report_type == 'all' or report_type == 'income':
            response_data['data']['income'] = get_income_report(start_date, end_date)

        if report_type == 'all' or report_type == 'expense':
            response_data['data']['expense'] = get_expense_report(start_date, end_date)

        if report_type == 'all' or report_type == 'purchase':
            response_data['data']['purchase'] = get_purchase_report(start_date, end_date)

        if report_type == 'all' or report_type == 'payroll':
            response_data['data']['payroll'] = get_payroll_report(start_date, end_date)

        if report_type == 'all' or report_type == 'attendance':
            response_data['data']['attendance'] = get_attendance_report(start_date, end_date)

        if report_type == 'all':
            response_data['data']['summary'] = get_overall_summary(
                response_data['data'], start_date, end_date
            )

        return JsonResponse(response_data)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


def get_income_report(start_date, end_date):
    """Generate income report data"""
    incomes = Income.objects.filter(
        date__range=[start_date, end_date]
    ).order_by('-date')

    total_income = incomes.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    income_count = incomes.count()
    days_in_period = (end_date - start_date).days + 1
    avg_daily = total_income / Decimal(str(days_in_period)) if days_in_period > 0 else Decimal('0')

    status_dist = incomes.values('status').annotate(
        count=Count('id'),
        amount=Sum('amount')
    )

    payment_dist = incomes.values('payment_mode').annotate(
        count=Count('id'),
        amount=Sum('amount')
    )

    return {
        'records': list(incomes.values(
            'id', 'date', 'description', 'payment_mode', 'amount', 'status'
        )),
        'summary': {
            'total': float(total_income),
            'count': income_count,
            'avg_daily': float(avg_daily),
            'status_distribution': list(status_dist),
            'payment_distribution': list(payment_dist)
        }
    }


def get_expense_report(start_date, end_date):
    """Generate expense report data"""
    expenses = Expense.objects.filter(
        date__range=[start_date, end_date],
        record_state='active'
    ).order_by('-date')

    total_expense = expenses.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    payroll_expenses = expenses.filter(category='Salary').aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0')
    other_expenses = total_expense - payroll_expenses

    category_dist = expenses.values('category').annotate(
        count=Count('id'),
        amount=Sum('total_amount')
    ).order_by('-amount')

    payment_dist = expenses.values('payment_method').annotate(
        count=Count('id'),
        amount=Sum('total_amount')
    )

    return {
        'records': list(expenses.values(
            'id', 'date', 'category', 'description',
            'total_amount', 'payment_method', 'employee_name'
        )),
        'summary': {
            'total': float(total_expense),
            'payroll': float(payroll_expenses),
            'other': float(other_expenses),
            'category_distribution': list(category_dist),
            'payment_distribution': list(payment_dist)
        }
    }


def get_purchase_report(start_date, end_date):
    """Generate purchase report data"""
    purchases = Purchase.objects.filter(
        date__range=[start_date, end_date]
    ).order_by('-date')

    total_purchase = purchases.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    purchase_count = purchases.count()
    pending_purchases = purchases.filter(status='Pending').aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0')

    vendor_dist = purchases.values('vendor').annotate(
        count=Count('id'),
        amount=Sum('total_amount')
    ).order_by('-amount')

    status_dist = purchases.values('status').annotate(
        count=Count('id'),
        amount=Sum('total_amount')
    )

    category_dist = purchases.values('cat__name').annotate(
        count=Count('id'),
        amount=Sum('total_amount')
    )

    return {
        'records': list(purchases.values(
            'id', 'date', 'vendor', 'cat__name', 'bill_no',
            'total_amount', 'status', 'description'
        )),
        'summary': {
            'total': float(total_purchase),
            'count': purchase_count,
            'pending': float(pending_purchases),
            'vendor_distribution': list(vendor_dist),
            'status_distribution': list(status_dist),
            'category_distribution': list(category_dist)
        }
    }


def get_payroll_report(start_date, end_date):
    """Generate payroll report data"""
    payrolls = Payroll.objects.filter(
        salary_date__range=[start_date, end_date],
        record_state='active'
    ).order_by('-salary_date')

    total_salary = payrolls.aggregate(total=Sum('net_salary'))['total'] or Decimal('0')
    total_incentives = payrolls.aggregate(total=Sum('spr_amount'))['total'] or Decimal('0')
    employees_paid = payrolls.values('employee_name').distinct().count()

    total_cash = payrolls.aggregate(total=Sum('cash_amount'))['total'] or Decimal('0')
    total_bank = payrolls.aggregate(total=Sum('bank_transfer_amount'))['total'] or Decimal('0')

    payment_dist = payrolls.values('payment_split_type').annotate(
        count=Count('id'),
        amount=Sum('net_salary')
    )

    employee_totals = payrolls.values('employee_name').annotate(
        total_salary=Sum('net_salary'),
        total_incentives=Sum('spr_amount'),
        count=Count('id')
    ).order_by('-total_salary')

    return {
        'records': list(payrolls.values(
            'id', 'employee_name', 'salary_date', 'month', 'year',
            'basic_pay', 'spr_amount', 'net_salary',
            'payment_split_type', 'cash_amount', 'bank_transfer_amount'
        )),
        'summary': {
            'total_salary': float(total_salary),
            'total_incentives': float(total_incentives),
            'employees_paid': employees_paid,
            'cash_total': float(total_cash),
            'bank_total': float(total_bank),
            'payment_distribution': list(payment_dist),
            'employee_totals': list(employee_totals)
        }
    }


def get_attendance_report(start_date, end_date):
    """Generate attendance report data"""
    summaries = AttendanceSummary.objects.filter(
        start_date__gte=start_date,
        end_date__lte=end_date,
        record_state='active',
        period_type='daily'
    )

    if not summaries.exists():
        summaries_data = []
        employees = Payroll.objects.filter(
            record_state='active'
        ).values_list('employee_name', flat=True).distinct()

        for employee in employees:
            summary, _ = AttendanceSummary.generate_summary_for_period(
                employee_name=employee,
                period_type='custom',
                start_date=start_date,
                end_date=end_date
            )
            summaries_data.append(summary)
    else:
        summaries_data = list(summaries)

    total_employees = len(summaries_data)
    total_present = sum(s.present_days for s in summaries_data)
    total_half = sum(s.half_days for s in summaries_data)
    total_absent = sum(s.absent_days for s in summaries_data)
    total_full_days = sum(float(s.full_days) for s in summaries_data)
    total_days_in_period = (end_date - start_date).days + 1

    avg_attendance = (total_full_days / (
        total_employees * total_days_in_period)) * 100 if total_employees > 0 and total_days_in_period > 0 else 0

    return {
        'records': [{
            'employee_name': s.employee_name,
            'present_days': s.present_days,
            'half_days': s.half_days,
            'absent_days': s.absent_days,
            'full_days': float(s.full_days),
            'total_days': s.total_days_in_period,
            'attendance_percentage': (
                float(s.full_days) / s.total_days_in_period * 100) if s.total_days_in_period > 0 else 0
        } for s in summaries_data],
        'summary': {
            'total_employees': total_employees,
            'total_present': total_present,
            'total_half': total_half,
            'total_absent': total_absent,
            'total_full_days': float(total_full_days),
            'total_days_in_period': total_days_in_period,
            'avg_attendance': float(avg_attendance)
        }
    }


def get_overall_summary(report_data, start_date, end_date):
    """Generate overall summary for combined report"""
    days_in_period = (end_date - start_date).days + 1

    total_income = report_data.get('income', {}).get('summary', {}).get('total', 0)
    total_expense = report_data.get('expense', {}).get('summary', {}).get('total', 0)
    total_purchase = report_data.get('purchase', {}).get('summary', {}).get('total', 0)
    total_salary = report_data.get('payroll', {}).get('summary', {}).get('total_salary', 0)

    net_profit = total_income - total_expense - total_purchase - total_salary

    return {
        'total_income': total_income,
        'total_expense': total_expense,
        'total_purchase': total_purchase,
        'total_salary': total_salary,
        'net_profit': net_profit,
        'days_in_period': days_in_period,
        'daily_avg_income': total_income / days_in_period if days_in_period > 0 else 0,
        'daily_avg_expense': (
            total_expense + total_purchase + total_salary) / days_in_period if days_in_period > 0 else 0
    }


@require_GET
def export_report(request):
    """Export report to Excel or PDF format"""
    try:
        format_type = request.GET.get('format', 'excel')
        report_type = request.GET.get('type', 'all')
        period_type = request.GET.get('period', 'daily')
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        # Get report data
        report_data = {}
        if report_type == 'all' or report_type == 'income':
            report_data['income'] = get_income_report(start_date, end_date)
        if report_type == 'all' or report_type == 'expense':
            report_data['expense'] = get_expense_report(start_date, end_date)
        if report_type == 'all' or report_type == 'purchase':
            report_data['purchase'] = get_purchase_report(start_date, end_date)
        if report_type == 'all' or report_type == 'payroll':
            report_data['payroll'] = get_payroll_report(start_date, end_date)
        if report_type == 'all' or report_type == 'attendance':
            report_data['attendance'] = get_attendance_report(start_date, end_date)

        if report_type == 'all':
            report_data['summary'] = get_overall_summary(report_data, start_date, end_date)

        if format_type == 'excel':
            return export_to_excel(report_data, report_type, start_date, end_date, period_type)
        elif format_type == 'pdf':
            return export_to_pdf(report_data, report_type, start_date, end_date, period_type)
        else:
            return JsonResponse({'success': False, 'error': 'Invalid format'}, status=400)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


def export_to_excel(report_data, report_type, start_date, end_date, period_type):
    """Export report to Excel format"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Styling
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def add_header(ws, title):
        """Add header to worksheet"""
        ws.append([title])
        ws.append([f"Period: {period_type.capitalize()}"])
        ws.append([f"Date Range: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}"])
        ws.append([])
        ws['A1'].font = title_font

    # Summary Sheet (if all reports)
    if report_type == 'all' and 'summary' in report_data:
        ws = wb.create_sheet('Summary')
        add_header(ws, 'Overall Financial Summary')
        
        summary = report_data['summary']
        ws.append(['Metric', 'Amount (₹)'])
        ws.append(['Total Income', summary['total_income']])
        ws.append(['Total Expenses', summary['total_expense']])
        ws.append(['Total Purchases', summary['total_purchase']])
        ws.append(['Total Salary', summary['total_salary']])
        ws.append(['Net Profit/Loss', summary['net_profit']])
        ws.append([])
        ws.append(['Days in Period', summary['days_in_period']])
        ws.append(['Daily Avg Income', summary['daily_avg_income']])
        ws.append(['Daily Avg Expense', summary['daily_avg_expense']])

        # Format header
        for cell in ws[5]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    # Income Sheet
    if 'income' in report_data:
        ws = wb.create_sheet('Income Report')
        add_header(ws, 'Income Report')
        
        # Headers
        headers = ['Date', 'Description', 'Payment Mode', 'Amount (₹)', 'Status']
        ws.append(headers)
        
        # Format headers
        for cell in ws[5]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for record in report_data['income']['records']:
            ws.append([
                record['date'].strftime('%d-%b-%Y') if isinstance(record['date'], datetime) else record['date'],
                record['description'],
                record['payment_mode'],
                record['amount'],
                record['status']
            ])
        
        # Summary
        ws.append([])
        summary = report_data['income']['summary']
        ws.append(['Total Income:', summary['total']])
        ws.append(['Total Records:', summary['count']])
        ws.append(['Daily Average:', summary['avg_daily']])

    # Expense Sheet
    if 'expense' in report_data:
        ws = wb.create_sheet('Expense Report')
        add_header(ws, 'Expense Report')
        
        headers = ['Date', 'Category', 'Description', 'Amount (₹)', 'Payment Method', 'Employee']
        ws.append(headers)
        
        for cell in ws[5]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for record in report_data['expense']['records']:
            ws.append([
                record['date'].strftime('%d-%b-%Y') if isinstance(record['date'], datetime) else record['date'],
                record['category'],
                record.get('description', '-'),
                record['total_amount'],
                record['payment_method'],
                record.get('employee_name', '-')
            ])
        
        ws.append([])
        summary = report_data['expense']['summary']
        ws.append(['Total Expenses:', summary['total']])
        ws.append(['Payroll Expenses:', summary['payroll']])
        ws.append(['Other Expenses:', summary['other']])

    # Purchase Sheet
    if 'purchase' in report_data:
        ws = wb.create_sheet('Purchase Report')
        add_header(ws, 'Purchase Report')
        
        headers = ['Date', 'Vendor', 'Category', 'Bill No', 'Amount (₹)', 'Status', 'Description']
        ws.append(headers)
        
        for cell in ws[5]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for record in report_data['purchase']['records']:
            ws.append([
                record['date'].strftime('%d-%b-%Y') if isinstance(record['date'], datetime) else record['date'],
                record['vendor'],
                record.get('cat__name', '-'),
                record['bill_no'],
                record['total_amount'],
                record['status'],
                record.get('description', '-')
            ])
        
        ws.append([])
        summary = report_data['purchase']['summary']
        ws.append(['Total Purchases:', summary['total']])
        ws.append(['Total Records:', summary['count']])
        ws.append(['Pending Amount:', summary['pending']])

    # Payroll Sheet
    if 'payroll' in report_data:
        ws = wb.create_sheet('Payroll Report')
        add_header(ws, 'Payroll Report')
        
        headers = ['Employee', 'Date', 'Month/Year', 'Basic Pay (₹)', 'Incentives (₹)', 'Net Salary (₹)', 'Payment Type']
        ws.append(headers)
        
        for cell in ws[5]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for record in report_data['payroll']['records']:
            ws.append([
                record['employee_name'],
                record['salary_date'].strftime('%d-%b-%Y') if isinstance(record['salary_date'], datetime) else record['salary_date'],
                f"{record['month']}/{record['year']}",
                record['basic_pay'],
                record['spr_amount'],
                record['net_salary'],
                record['payment_split_type'].replace('_', ' ').title()
            ])
        
        ws.append([])
        summary = report_data['payroll']['summary']
        ws.append(['Total Salary:', summary['total_salary']])
        ws.append(['Total Incentives:', summary['total_incentives']])
        ws.append(['Employees Paid:', summary['employees_paid']])
        ws.append(['Cash Payments:', summary['cash_total']])
        ws.append(['Bank Payments:', summary['bank_total']])

    # Attendance Sheet
    if 'attendance' in report_data:
        ws = wb.create_sheet('Attendance Report')
        add_header(ws, 'Attendance Report')
        
        headers = ['Employee', 'Present Days', 'Half Days', 'Absent Days', 'Full Day Equiv.', 'Attendance %']
        ws.append(headers)
        
        for cell in ws[5]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for record in report_data['attendance']['records']:
            ws.append([
                record['employee_name'],
                record['present_days'],
                record['half_days'],
                record['absent_days'],
                record['full_days'],
                f"{record['attendance_percentage']:.2f}%"
            ])
        
        ws.append([])
        summary = report_data['attendance']['summary']
        ws.append(['Total Employees:', summary['total_employees']])
        ws.append(['Total Present:', summary['total_present']])
        ws.append(['Total Absent:', summary['total_absent']])
        ws.append(['Average Attendance:', f"{summary['avg_attendance']:.2f}%"])

    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Financial_Report_{report_type}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def export_to_pdf(report_data, report_type, start_date, end_date, period_type):
    """Export report to PDF format using ReportLab"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from io import BytesIO
    
    # Create PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Container for PDF elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#374151'),
        spaceAfter=10,
        spaceBefore=15
    )
    
    # Title
    elements.append(Paragraph('Financial Report', title_style))
    elements.append(Paragraph(f"Period: {period_type.capitalize()}", styles['Normal']))
    elements.append(Paragraph(f"Date Range: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Summary Section (if all reports)
    if report_type == 'all' and 'summary' in report_data:
        summary = report_data['summary']
        elements.append(Paragraph('Overall Financial Summary', heading_style))
        
        summary_data = [
            ['Metric', 'Amount (₹)'],
            ['Total Income', f"₹{summary['total_income']:,.2f}"],
            ['Total Expenses', f"₹{summary['total_expense']:,.2f}"],
            ['Total Purchases', f"₹{summary['total_purchase']:,.2f}"],
            ['Total Salary', f"₹{summary['total_salary']:,.2f}"],
            ['Net Profit/Loss', f"₹{summary['net_profit']:,.2f}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
    
    # Income Report
    if 'income' in report_data:
        elements.append(Paragraph('Income Report', heading_style))
        
        income_data = [['Date', 'Description', 'Payment Mode', 'Amount (₹)', 'Status']]
        for record in report_data['income']['records'][:50]:  # Limit to 50 records
            income_data.append([
                record['date'][:10] if isinstance(record['date'], str) else str(record['date']),
                str(record['description'])[:30],
                record['payment_mode'],
                f"₹{record['amount']:,.2f}",
                record['status']
            ])
        
        if report_data['income']['records']:
            income_table = Table(income_data, colWidths=[1*inch, 2*inch, 1.2*inch, 1.2*inch, 0.8*inch])
            income_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(income_table)
            
            # Summary
            summary = report_data['income']['summary']
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(f"<b>Total Income:</b> ₹{summary['total']:,.2f} | <b>Count:</b> {summary['count']}", styles['Normal']))
        elements.append(PageBreak())
    
    # Expense Report
    if 'expense' in report_data:
        elements.append(Paragraph('Expense Report', heading_style))
        
        expense_data = [['Date', 'Category', 'Description', 'Amount (₹)', 'Payment']]
        for record in report_data['expense']['records'][:50]:
            expense_data.append([
                record['date'][:10] if isinstance(record['date'], str) else str(record['date']),
                record['category'][:15],
                str(record.get('description', '-'))[:25],
                f"₹{record['total_amount']:,.2f}",
                record['payment_method'][:10]
            ])
        
        if report_data['expense']['records']:
            expense_table = Table(expense_data, colWidths=[1*inch, 1.2*inch, 2*inch, 1.2*inch, 1*inch])
            expense_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(expense_table)
            
            summary = report_data['expense']['summary']
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(f"<b>Total Expenses:</b> ₹{summary['total']:,.2f}", styles['Normal']))
        elements.append(PageBreak())
    
    # Purchase Report
    if 'purchase' in report_data:
        elements.append(Paragraph('Purchase Report', heading_style))
        
        purchase_data = [['Date', 'Vendor', 'Category', 'Bill No', 'Amount (₹)', 'Status']]
        for record in report_data['purchase']['records'][:50]:
            purchase_data.append([
                record['date'][:10] if isinstance(record['date'], str) else str(record['date']),
                str(record['vendor'])[:20],
                str(record.get('cat__name', '-'))[:15],
                str(record['bill_no']),
                f"₹{record['total_amount']:,.2f}",
                record['status']
            ])
        
        if report_data['purchase']['records']:
            purchase_table = Table(purchase_data, colWidths=[0.9*inch, 1.5*inch, 1*inch, 0.8*inch, 1.2*inch, 0.8*inch])
            purchase_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(purchase_table)
            
            summary = report_data['purchase']['summary']
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(f"<b>Total Purchases:</b> ₹{summary['total']:,.2f} | <b>Pending:</b> ₹{summary['pending']:,.2f}", styles['Normal']))
        elements.append(PageBreak())
    
    # Payroll Report
    if 'payroll' in report_data:
        elements.append(Paragraph('Payroll Report', heading_style))
        
        payroll_data = [['Employee', 'Date', 'Basic (₹)', 'Incentives (₹)', 'Net Salary (₹)']]
        for record in report_data['payroll']['records'][:50]:
            payroll_data.append([
                str(record['employee_name'])[:25],
                record['salary_date'][:10] if isinstance(record['salary_date'], str) else str(record['salary_date']),
                f"₹{record['basic_pay']:,.2f}",
                f"₹{record['spr_amount']:,.2f}",
                f"₹{record['net_salary']:,.2f}"
            ])
        
        if report_data['payroll']['records']:
            payroll_table = Table(payroll_data, colWidths=[2*inch, 1*inch, 1.2*inch, 1.2*inch, 1.2*inch])
            payroll_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(payroll_table)
            
            summary = report_data['payroll']['summary']
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(f"<b>Total Salary:</b> ₹{summary['total_salary']:,.2f} | <b>Employees:</b> {summary['employees_paid']}", styles['Normal']))
        elements.append(PageBreak())
    
    # Attendance Report
    if 'attendance' in report_data:
        elements.append(Paragraph('Attendance Report', heading_style))
        
        attendance_data = [['Employee', 'Present', 'Half Day', 'Absent', 'Full Days', 'Attendance %']]
        for record in report_data['attendance']['records'][:50]:
            attendance_data.append([
                str(record['employee_name'])[:30],
                str(record['present_days']),
                str(record['half_days']),
                str(record['absent_days']),
                f"{record['full_days']:.1f}",
                f"{record['attendance_percentage']:.1f}%"
            ])
        
        if report_data['attendance']['records']:
            attendance_table = Table(attendance_data, colWidths=[2.5*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1*inch])
            attendance_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(attendance_table)
            
            summary = report_data['attendance']['summary']
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(f"<b>Total Employees:</b> {summary['total_employees']} | <b>Avg Attendance:</b> {summary['avg_attendance']:.1f}%", styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    
    # Create response
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f"Financial_Report_{report_type}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# Legacy functions
@require_GET
def get_combined_report(request):
    """Legacy endpoint for combined report"""
    return get_report_data(request)


@require_GET
def get_attendance_summary_report(request):
    """Legacy endpoint for attendance summary report"""
    try:
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        if not start_date_str or not end_date_str:
            return JsonResponse({
                'success': False,
                'error': 'Start date and end date required'
            }, status=400)

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        attendance_data = get_attendance_report(start_date, end_date)

        return JsonResponse({
            'success': True,
            'data': attendance_data
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)